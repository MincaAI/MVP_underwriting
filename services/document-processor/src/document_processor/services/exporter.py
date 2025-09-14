"""
Document Exporter Service

Handles Excel export generation with professional formatting.
Migrated from worker-exporter with Gcotiza template support.
"""

import sys
import pathlib
import uuid
import io
import yaml
import pandas as pd
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime

# Add packages to path
sys.path.insert(0, str(pathlib.Path(__file__).parent.parent.parent.parent.parent.parent / "packages" / "db" / "src"))
sys.path.insert(0, str(pathlib.Path(__file__).parent.parent.parent.parent.parent.parent / "packages" / "storage" / "src"))

from sqlalchemy.orm import Session
from app.db.session import engine
from app.db.models import Run, Transform, Codify, Export, RunStatus, Component
from app.storage.s3 import upload_bytes
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class DocumentExporter:
    """
    Exports processed data to Excel format with professional styling.
    Supports Gcotiza templates and custom export configurations.
    """
    
    def __init__(self):
        self.logger = logger
        self.templates_path = pathlib.Path(__file__).parent.parent.parent.parent.parent.parent / "configs" / "gcotiza-templates"
    
    async def export(self, run_id: str, template: str = "gcotiza_v1.yaml") -> str:
        """
        Export processed data to Excel format.
        
        Args:
            run_id: Base run identifier
            template: Export template name
            
        Returns:
            S3 URL of the generated Excel file
        """
        try:
            # Create export run record
            export_run_id = f"{run_id}_export"
            with Session(engine) as session:
                export_run = Run(
                    id=export_run_id,
                    case_id=session.get(Run, run_id).case_id if session.get(Run, run_id) else "",
                    component=Component.EXPORT,
                    status=RunStatus.STARTED,
                    parent_run_id=run_id,
                    started_at=datetime.utcnow()
                )
                session.add(export_run)
                session.commit()
            
            # Load export template
            template_config = await self._load_template(template)
            
            # Get processed data
            data = await self._get_processed_data(run_id)
            
            if not data:
                raise ValueError(f"No processed data found for run {run_id}")
            
            # Create Excel workbook
            excel_bytes = await self._create_excel(data, template_config)
            
            # Generate filename and upload to S3
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{run_id}_{timestamp}.xlsx"
            s3_key = f"exports/{run_id}/{filename}"
            
            export_url = upload_bytes(excel_bytes, s3_key)
            
            # Store export record
            await self._store_export_record(export_run_id, run_id, export_url, template_config, len(data))
            
            # Update export run with completion
            with Session(engine) as session:
                export_run = session.get(Run, export_run_id)
                export_run.status = RunStatus.COMPLETED
                export_run.completed_at = datetime.utcnow()
                export_run.metrics = {
                    'rows_exported': len(data),
                    'template_used': template,
                    'file_size': len(excel_bytes),
                    'export_url': export_url
                }
                session.commit()
            
            self.logger.info(f"Successfully exported {len(data)} rows to {export_url}")
            return export_url
            
        except Exception as e:
            # Update run with error status
            with Session(engine) as session:
                export_run = session.get(Run, export_run_id)
                if export_run:
                    export_run.status = RunStatus.ERROR
                    export_run.error_message = str(e)
                    session.commit()
            
            self.logger.error(f"Error exporting data for run {run_id}: {e}")
            raise
    
    async def _load_template(self, template: str) -> Dict[str, Any]:
        """Load export template configuration"""
        try:
            template_path = self.templates_path / template
            
            if not template_path.exists():
                # Return default Gcotiza template
                return self._get_default_gcotiza_template()
            
            with open(template_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f)
            
            return config
            
        except Exception as e:
            self.logger.warning(f"Error loading template {template}: {e}. Using default.")
            return self._get_default_gcotiza_template()
    
    def _get_default_gcotiza_template(self) -> Dict[str, Any]:
        """Get default Gcotiza export template"""
        return {
            'name': 'Gcotiza v1 Template',
            'description': 'Standard Gcotiza export format with 46 columns',
            'columns': [
                {'name': 'POLIZA', 'source': 'policy_number', 'type': 'text'},
                {'name': 'CERTIFICADO', 'source': 'certificate', 'type': 'text'},
                {'name': 'ITEM', 'source': 'item', 'type': 'int'},
                {'name': 'PLACA', 'source': 'license_plate', 'type': 'text'},
                {'name': 'VIN', 'source': 'vin', 'type': 'text'},
                {'name': 'MARCA', 'source': 'brand', 'type': 'text'},
                {'name': 'MODELO', 'source': 'model', 'type': 'text'},
                {'name': 'AÑO', 'source': 'model_year', 'type': 'int'},
                {'name': 'DESCRIPCION', 'source': 'description', 'type': 'text'},
                {'name': 'COLOR', 'source': 'color', 'type': 'text'},
                {'name': 'TIPO_COMBUSTIBLE', 'source': 'fuel_type', 'type': 'text'},
                {'name': 'SUMA_ASEGURADA', 'source': 'insured_value', 'type': 'float'},
                {'name': 'PRIMA', 'source': 'premium', 'type': 'float'},
                {'name': 'DEDUCIBLE', 'source': 'deductible', 'type': 'float'},
                {'name': 'COBERTURA', 'source': 'coverage_type', 'type': 'text'},
                # Additional Gcotiza columns...
                {'name': 'CLAVE_VEHICULAR', 'source': 'cvegs_code', 'type': 'text'},
                {'name': 'CONFIANZA_IA', 'source': 'confidence', 'type': 'percent'},
                {'name': 'ESTADO_PROCESO', 'source': 'process_status', 'type': 'text'},
                {'name': 'FECHA_PROCESO', 'source': 'processed_at', 'type': 'date'}
            ],
            'formatting': {
                'freeze_header': True,
                'header_style': {
                    'font_color': 'FFFFFF',
                    'bg_color': '4472C4',
                    'bold': True
                },
                'data_style': {
                    'font_color': '000000',
                    'bg_color': 'FFFFFF'
                }
            }
        }
    
    async def _get_processed_data(self, run_id: str) -> List[Dict[str, Any]]:
        """Get processed data from transform and codify results"""
        try:
            processed_data = []
            
            with Session(engine) as session:
                # Get transform run
                transform_run_id = f"{run_id}_transform"
                transforms = session.query(Transform).filter(Transform.run_id == transform_run_id).all()
                
                # Get codify run if exists
                codify_run_id = f"{run_id}_codify"
                codifies = {c.row_id: c for c in session.query(Codify).filter(Codify.run_id == codify_run_id).all()}
                
                for transform in transforms:
                    # Merge transform and codify data
                    row_data = transform.transformed_data.copy() if transform.transformed_data else {}
                    
                    # Add codification results if available
                    codify = codifies.get(transform.row_id)
                    if codify and codify.result:
                        row_data.update({
                            'cvegs_code': codify.result.get('cvegs_code'),
                            'confidence': codify.result.get('confidence'),
                            'codify_status': codify.result.get('status', 'completed')
                        })
                    else:
                        row_data.update({
                            'cvegs_code': None,
                            'confidence': None,
                            'codify_status': 'not_processed'
                        })
                    
                    # Add processing metadata
                    row_data.update({
                        'process_status': 'completed',
                        'processed_at': datetime.utcnow(),
                        'transform_errors': transform.validation_errors or []
                    })
                    
                    processed_data.append(row_data)
            
            return processed_data
            
        except Exception as e:
            self.logger.error(f"Error getting processed data for run {run_id}: {e}")
            raise
    
    async def _create_excel(self, data: List[Dict[str, Any]], template_config: Dict[str, Any]) -> bytes:
        """Create Excel file from processed data"""
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Gcotiza Export"
            
            # Get column configuration
            columns = template_config.get('columns', [])
            formatting = template_config.get('formatting', {})
            
            # Create DataFrame with proper column mapping
            df_data = []
            for row_data in data:
                excel_row = {}
                for col_config in columns:
                    col_name = col_config['name']
                    source_field = col_config['source']
                    col_type = col_config.get('type', 'text')
                    
                    # Get value and apply type conversion
                    value = self._get_nested_value(row_data, source_field)
                    converted_value = self._convert_value(value, col_type)
                    excel_row[col_name] = converted_value
                
                df_data.append(excel_row)
            
            # Create DataFrame
            df = pd.DataFrame(df_data)
            
            # Write data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Apply formatting
            await self._apply_excel_formatting(ws, formatting, len(columns), len(data) + 1)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            self.logger.error(f"Error creating Excel file: {e}")
            raise
    
    def _get_nested_value(self, data: Dict[str, Any], key: str) -> Any:
        """Get nested dictionary value using dot notation"""
        keys = key.split(".")
        value = data
        
        for k in keys:
            if isinstance(value, dict) and k in value:
                value = value[k]
            else:
                return None
        
        return value
    
    def _convert_value(self, value: Any, value_type: str) -> Any:
        """Convert value to specified type for Excel"""
        if value is None or value == "":
            return None
        
        try:
            if value_type == "int":
                return int(float(str(value))) if str(value).replace('.', '').replace('-', '').isdigit() else None
            elif value_type == "float":
                return float(value)
            elif value_type == "percent":
                # Convert to decimal (0.05 for 5%)
                str_value = str(value).replace('%', '')
                return float(str_value) / 100.0
            elif value_type == "date":
                if isinstance(value, datetime):
                    return value
                return pd.to_datetime(value, errors='coerce')
            else:  # text
                return str(value) if value is not None else ""
        except (ValueError, TypeError):
            return value  # Return original value if conversion fails
    
    async def _apply_excel_formatting(self, worksheet, formatting: Dict[str, Any], num_cols: int, num_rows: int):
        """Apply professional formatting to Excel worksheet"""
        try:
            # Header formatting
            header_style = formatting.get('header_style', {})
            if header_style:
                header_font = Font(
                    color=header_style.get('font_color', 'FFFFFF'),
                    bold=header_style.get('bold', True)
                )
                header_fill = PatternFill(
                    start_color=header_style.get('bg_color', '4472C4'),
                    end_color=header_style.get('bg_color', '4472C4'),
                    fill_type='solid'
                )
                
                # Apply to header row
                for col in range(1, num_cols + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data formatting
            data_style = formatting.get('data_style', {})
            if data_style:
                data_font = Font(color=data_style.get('font_color', '000000'))
                
                # Apply to data rows
                for row in range(2, num_rows + 1):
                    for col in range(1, num_cols + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = data_font
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, num_rows + 1):
                for col in range(1, num_cols + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Max width of 50
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row if specified
            if formatting.get('freeze_header'):
                worksheet.freeze_panes = 'A2'
                
        except Exception as e:
            self.logger.warning(f"Error applying Excel formatting: {e}")
    
    async def _store_export_record(self, export_run_id: str, base_run_id: str, export_url: str, 
                                 template_config: Dict[str, Any], row_count: int):
        """Store export record in database"""
        try:
            with Session(engine) as session:
                export = Export(
                    id=str(uuid.uuid4()),
                    run_id=export_run_id,
                    base_run_id=base_run_id,
                    export_url=export_url,
                    template_name=template_config.get('name', 'Unknown'),
                    row_count=row_count,
                    created_at=datetime.utcnow()
                )
                session.add(export)
                session.commit()
                
                self.logger.info(f"Stored export record for run {export_run_id}")
                
        except Exception as e:
            self.logger.error(f"Error storing export record: {e}")
            raise